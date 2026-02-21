from flask import Flask, request, jsonify, send_file, render_template
import sqlite3, os, io, datetime, json, threading, time, signal
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment

app = Flask(__name__)

# --- OTOMATİK KAPANMA (HEARTBEAT) ---
last_received_heartbeat = time.time()

def shutdown_server():
    print("Tarayıcı kapatıldı, uygulama sonlandırılıyor...")
    os.kill(os.getpid(), signal.SIGTERM)

def check_heartbeat():
    global last_received_heartbeat
    while True:
        time.sleep(10)
        # 25 saniye boyunca tarayıcıdan sinyal gelmezse uygulamayı kapatır
        if time.time() - last_received_heartbeat > 25:
            shutdown_server()

threading.Thread(target=check_heartbeat, daemon=True).start()

@app.route('/heartbeat', methods=['POST'])
def heartbeat():
    global last_received_heartbeat
    last_received_heartbeat = time.time()
    return "ok"

# --- VERİTABANI AYARLARI ---
APP_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(APP_DIR, "data.db")

STATUSES = ["DEĞERLENDİRME","EKSİK","ONAY","RET","İPTAL"]
STATUS_BG = {"DEĞERLENDİRME":"#f8f8ff","EKSİK":"#528b8b","ONAY":"#f5f5dc","RET":"#c62828","İPTAL":"#8b3e2f"}
STATUS_FG = {"DEĞERLENDİRME":"#222222","EKSİK":"#ffffff","ONAY":"#222222","RET":"#ffffff","İPTAL":"#ffffff"}

def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = db(); cur = conn.cursor()
    cur.execute("""CREATE TABLE IF NOT EXISTS calisma_izni(
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      tarih TEXT, firma TEXT, basvuru_no TEXT, yabanci_isim TEXT, kimlik_no TEXT, ulke TEXT,
      durum TEXT, notlar TEXT, created_at TEXT)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS ikamet_izni(
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      tarih TEXT, ad_soyad TEXT, referans TEXT, iletisim TEXT, pasaport_tc TEXT,
      durum TEXT, notlar TEXT, created_at TEXT)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS sgk(
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      tarih TEXT, isveren_isim TEXT, tc TEXT, dosya_no TEXT, sistem_sifre TEXT, isyeri_sifre TEXT,
      notlar TEXT, created_at TEXT)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS eimza(
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      tarih TEXT, musteri_adi TEXT, suresi TEXT, odeme TEXT, kargo_durumu TEXT,
      notlar TEXT, created_at TEXT)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS reminders(
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      tarih TEXT, baslik TEXT, aciklama TEXT, created_at TEXT)""")
    
    try:
        cols = [r[1] for r in cur.execute("PRAGMA table_info(sgk)").fetchall()]
        if "dosya_no" not in cols:
            cur.execute("ALTER TABLE sgk ADD COLUMN dosya_no TEXT DEFAULT ''")
    except Exception:
        pass

    conn.commit(); conn.close()

def parse_date_any(v):
    if v is None: return None
    if isinstance(v, (datetime.date, datetime.datetime)):
        d = v.date() if isinstance(v, datetime.datetime) else v
        return d.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s: return None
    fmts = ["%d.%m.%Y","%d.%m.%y","%d/%m/%Y","%d/%m/%y","%Y-%m-%d","%Y/%m/%d"]
    for f in fmts:
        try: return datetime.datetime.strptime(s, f).date().strftime("%Y-%m-%d")
        except: pass
    try:
        x = float(s)
        if 20000 < x < 60000:
            base = datetime.date(1899, 12, 30)
            return (base + datetime.timedelta(days=int(x))).strftime("%Y-%m-%d")
    except: pass
    return None

def iso_to_tr(iso):
    if not iso: return ""
    try: return datetime.datetime.strptime(iso, "%Y-%m-%d").strftime("%d.%m.%Y")
    except: return iso

def status_from_cell(v):
    if v is None: return "DEĞERLENDİRME"
    s = str(v).strip().upper()
    if not s: return "DEĞERLENDİRME"
    if s == "ONAY": return "ONAY"
    if s == "RET": return "RET"
    for st in STATUSES:
        if s.replace("I","İ") == st: return st
    return "DEĞERLENDİRME"

def infer_headers(row):
    def n(s):
        s = str(s or "").strip().upper().replace("İ","I")
        return " ".join(s.split())
    return [n(c) for c in row]

def excel_export(table):
    wb = Workbook(); ws = wb.active
    conn = db(); cur = conn.cursor()

    if table == "calisma_izni":
        ws.title = "Çalışma İzni"
        ws.append(["TARİH","FİRMA / MÜŞTERİ","BAŞVURU NO","YABANCI İSİM","KİMLİK NO","ÜLKE","DURUM"])
        rows = cur.execute("SELECT tarih,firma,basvuru_no,yabanci_isim,kimlik_no,ulke,durum FROM calisma_izni ORDER BY (tarih IS NULL), tarih DESC, id DESC").fetchall()
        for r in rows: ws.append([iso_to_tr(r["tarih"]), r["firma"], r["basvuru_no"], r["yabanci_isim"], r["kimlik_no"], r["ulke"], r["durum"]])
        status_col = 7
    elif table == "ikamet_izni":
        ws.title = "İkamet İzni"
        ws.append(["TARİH","AD SOYAD","REFERANS","İLETİŞİM","PASAPORT / T.C.","DURUM"])
        rows = cur.execute("SELECT tarih,ad_soyad,referans,iletisim,pasaport_tc,durum FROM ikamet_izni ORDER BY (tarih IS NULL), tarih DESC, id DESC").fetchall()
        for r in rows: ws.append([iso_to_tr(r["tarih"]), r["ad_soyad"], r["referans"], r["iletisim"], r["pasaport_tc"], r["durum"]])
        status_col = 6
    elif table == "sgk":
        ws.title = "SGK"
        ws.append(["TARİH","İŞVEREN İSİM","T.C.","#","SİSTEM Ş.","İŞYERİ Ş."])
        rows = cur.execute("SELECT tarih,isveren_isim,tc,dosya_no,sistem_sifre,isyeri_sifre FROM sgk ORDER BY (tarih IS NULL), tarih DESC, id DESC").fetchall()
        for r in rows: ws.append([iso_to_tr(r["tarih"]), r["isveren_isim"], r["tc"], r["dosya_no"], r["sistem_sifre"], r["isyeri_sifre"]])
        status_col = None
    else:
        ws.title = "E-İmza"
        ws.append(["TARİH","MÜŞTERİ ADI","SÜRESİ","ÖDEME","KARGO DURUMU"])
        rows = cur.execute("SELECT tarih,musteri_adi,suresi,odeme,kargo_durumu FROM eimza ORDER BY (tarih IS NULL), tarih DESC, id DESC").fetchall()
        for r in rows: ws.append([iso_to_tr(r["tarih"]), r["musteri_adi"], r["suresi"], r["odeme"], r["kargo_durumu"]])
        status_col = None

    conn.close()

    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)
    for c in ws[1]:
        c.fill = header_fill; c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws.columns:
        mx = 10
        letter = col[0].column_letter
        for cell in col:
            v = cell.value
            mx = max(mx, len(str(v)) if v is not None else 0)
        ws.column_dimensions[letter].width = min(40, mx+2)

    if status_col:
        for r in range(2, ws.max_row+1):
            st = ws.cell(r, status_col).value
            bg = STATUS_BG.get(st, "#ffffff").replace("#","").upper()
            fg = STATUS_FG.get(st, "#222222").replace("#","").upper()
            fill = PatternFill("solid", fgColor=bg)
            font = Font(color=fg)
            for c in range(1, ws.max_column+1):
                ws.cell(r,c).fill = fill
                ws.cell(r,c).font = font

    bio = io.BytesIO()
    wb.save(bio); bio.seek(0)
    return bio

def import_xlsx(file_storage, target):
    wb = load_workbook(file_storage, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return 0
    headers = infer_headers(rows[0])

    def col(*names):
        for nm in names:
            nm2 = nm.upper().replace("İ","I")
            if nm2 in headers: return headers.index(nm2)
        return None

    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = db(); cur = conn.cursor()
    inserted = 0

    if target == "calisma_izni":
        i_firma = col("FIRMA","MUSTERI","MUSTERI ADI","FIRMA / MUSTERI","FİRMA / MÜŞTERİ")
        i_basv  = col("BASVURU NO","BAŞVURU NO","BASVURU NUMARASI")
        i_kisi  = col("KISI","KİŞİ","YABANCI ISIM","YABANCI İSİM","AD SOYAD")
        i_ulke  = col("ULKE","ÜLKE")
        i_tarih = col("TARIH","TARİH","BASVURU TAR","BAŞVURU TAR","BASVURU TARIH","BAŞVURU TARİH")
        i_durum = col("DURUM","ONAY")
        i_kimlik = col("KIMLIK NO","KİMLİK NO","T.C.","TC","PASAPORT","PASAPORT / T.C.","YABANCI NO")
        if i_kimlik is None:
            best_i, best_s = None, 0
            for idx in range(len(headers)):
                if idx == i_tarih: continue
                s = 0
                for rr in rows[1:101]:
                    v = rr[idx] if idx < len(rr) else None
                    if v is None: continue
                    ss = str(v).strip()
                    if ss and sum(ch.isdigit() for ch in ss) >= max(6, int(len(ss)*0.7)):
                        s += 1
                if s > best_s: best_s, best_i = s, idx
            i_kimlik = best_i

        for rr in rows[1:]:
            tarih = parse_date_any(rr[i_tarih] if i_tarih is not None and i_tarih < len(rr) else None)
            firma = str(rr[i_firma]).strip() if i_firma is not None and i_firma < len(rr) and rr[i_firma] is not None else ""
            basv  = str(rr[i_basv]).strip() if i_basv is not None and i_basv < len(rr) and rr[i_basv] is not None else ""
            kisi  = str(rr[i_kisi]).strip() if i_kisi is not None and i_kisi < len(rr) and rr[i_kisi] is not None else ""
            ulke  = str(rr[i_ulke]).strip() if i_ulke is not None and i_ulke < len(rr) and rr[i_ulke] is not None else ""
            kimlik= str(rr[i_kimlik]).strip() if i_kimlik is not None and i_kimlik < len(rr) and rr[i_kimlik] is not None else ""
            durum = status_from_cell(rr[i_durum] if i_durum is not None and i_durum < len(rr) else None)
            if not any([tarih,firma,basv,kisi,ulke,kimlik,durum]): continue
            cur.execute("INSERT INTO calisma_izni(tarih,firma,basvuru_no,yabanci_isim,kimlik_no,ulke,durum,notlar,created_at) VALUES(?,?,?,?,?,?,?,?,?)",
                        (tarih,firma,basv,kisi,kimlik,ulke,durum,"",now))
            inserted += 1

    elif target == "ikamet_izni":
        i_ad   = col("AD SOYAD","YABANCI ISIM","YABANCI İSİM")
        i_ref  = col("REFERANS","REFARANS","REF")
        i_ilet = col("ILETISIM","İLETİŞİM","EMAIL","E-MAIL","MAIL")
        i_pas  = col("PASAPORT / T.C.","PASAPORT","T.C.","TC")
        i_tarih= col("TARIH","TARİH")
        i_durum= col("DURUM","ONAY")
        for rr in rows[1:]:
            tarih = parse_date_any(rr[i_tarih] if i_tarih is not None and i_tarih < len(rr) else None)
            ad    = str(rr[i_ad]).strip() if i_ad is not None and i_ad < len(rr) and rr[i_ad] is not None else ""
            ref   = str(rr[i_ref]).strip() if i_ref is not None and i_ref < len(rr) and rr[i_ref] is not None else ""
            ilet  = str(rr[i_ilet]).strip() if i_ilet is not None and i_ilet < len(rr) and rr[i_ilet] is not None else ""
            pas   = str(rr[i_pas]).strip() if i_pas is not None and i_pas < len(rr) and rr[i_pas] is not None else ""
            durum = status_from_cell(rr[i_durum] if i_durum is not None and i_durum < len(rr) else None)
            if not any([tarih,ad,ref,ilet,pas,durum]): continue
            cur.execute("INSERT INTO ikamet_izni(tarih,ad_soyad,referans,iletisim,pasaport_tc,durum,notlar,created_at) VALUES(?,?,?,?,?,?,?,?)",
                        (tarih,ad,ref,ilet,pas,durum,"",now))
            inserted += 1

    elif target == "sgk":
        i_tarih = col("TARIH", "TARİH")
        i_isv   = col("ISVEREN ISIM", "İŞVEREN İSİM")
        i_tc    = col("T.C.", "TC")
        i_no    = col("#", "DOSYA NO", "NO") # Excel'deki başlığı yakalar
        i_sys   = col("SISTEM Ş.", "SİSTEM Ş.")
        i_isy   = col("ISYERI Ş.", "İŞYERİ Ş.")
        
        for rr in rows[1:]:
            if not any(rr): continue
            tarih = parse_date_any(rr[i_tarih] if i_tarih is not None else None)
            
            # Veriyi güvenli bir şekilde metne çevirip alıyoruz
            dosya_no_verisi = str(rr[i_no]).strip() if i_no is not None and rr[i_no] is not None else ""
            
            cur.execute("""INSERT INTO sgk(tarih, isveren_isim, tc, dosya_no, sistem_sifre, isyeri_sifre, notlar, created_at) 
                           VALUES(?,?,?,?,?,?,?,?)""",
                        (tarih, 
                         str(rr[i_isv] or ""), 
                         str(rr[i_tc] or ""), 
                         dosya_no_verisi, # Burası '#' sütunundaki veriyi veritabanına yazar
                         str(rr[i_sys] or ""), 
                         str(rr[i_isy] or ""), 
                         "Excel'den aktarıldı", 
                         now))
            inserted += 1

    else:
        i_tarih = col("TARIH","TARİH")
        i_mus   = col("MUSTERI ADI","MÜŞTERİ ADI","MUSTERI","MÜŞTERİ")
        i_sur   = col("SURESI","SÜRESİ","SURE","SÜRE")
        i_ode   = col("ODEME","ÖDEME")
        i_kar   = col("KARGO DURUMU","KARGO")
        for rr in rows[1:]:
            tarih = parse_date_any(rr[i_tarih] if i_tarih is not None and i_tarih < len(rr) else None)
            mus   = str(rr[i_mus]).strip() if i_mus is not None and i_mus < len(rr) and rr[i_mus] is not None else ""
            sur   = str(rr[i_sur]).strip() if i_sur is not None and i_sur < len(rr) and rr[i_sur] is not None else ""
            ode   = str(rr[i_ode]).strip() if i_ode is not None and i_ode < len(rr) and rr[i_ode] is not None else ""
            kar   = str(rr[i_kar]).strip() if i_kar is not None and i_kar < len(rr) and rr[i_kar] is not None else ""
            if not any([tarih,mus,sur,ode,kar]): continue
            cur.execute("INSERT INTO eimza(tarih,musteri_adi,suresi,odeme,kargo_durumu,notlar,created_at) VALUES(?,?,?,?,?,?,?)",
                        (tarih,mus,sur,ode,kar,"",now))
            inserted += 1

    conn.commit(); conn.close()
    return inserted

@app.route("/")
def home():
    return render_template("index.html")

@app.route("/api/meta")
def meta():
    return jsonify({"statuses": STATUSES, "status_bg": STATUS_BG, "status_fg": STATUS_FG})

@app.route("/api/stats")
def stats():
    conn = db(); cur = conn.cursor()
    def count_table(t):
        out = {}
        if t in ("calisma_izni","ikamet_izni"):
            for st in STATUSES:
                out[st] = cur.execute(f"SELECT COUNT(*) c FROM {t} WHERE durum=?", (st,)).fetchone()["c"]
        else:
            out["TOPLAM"] = cur.execute(f"SELECT COUNT(*) c FROM {t}").fetchone()["c"]
        return out
    data = {"calisma": count_table("calisma_izni"), "ikamet": count_table("ikamet_izni"),
            "sgk": count_table("sgk"), "eimza": count_table("eimza")}
    conn.close(); return jsonify(data)

@app.route("/api/recent")
def recent():
    conn = db(); cur = conn.cursor()
    rows = cur.execute("""
      SELECT 'Çalışma İzni' AS tip, tarih, firma AS ana, basvuru_no AS no, yabanci_isim AS isim, ulke, durum FROM calisma_izni
      UNION ALL
      SELECT 'İkamet İzni' AS tip, tarih, '' AS ana, referans AS no, ad_soyad AS isim, '' AS ulke, durum FROM ikamet_izni
      ORDER BY (tarih IS NULL), tarih DESC LIMIT 20
    """).fetchall()
    conn.close(); return jsonify([dict(r) for r in rows])

@app.route("/api/list/<table>")
def list_table(table):
    allowed = {"calisma_izni","ikamet_izni","sgk","eimza"}
    if table not in allowed: return jsonify({"error":"invalid"}), 400
    q = (request.args.get("q") or "").strip().lower()
    statuses = request.args.get("statuses")
    status_list = [s for s in statuses.split(",") if s] if statuses else None
    conn = db(); cur = conn.cursor()

    if table == "calisma_izni":
        sql = "SELECT * FROM calisma_izni WHERE 1=1"; params=[]
        if q:
            sql += " AND (lower(ifnull(firma,'')) LIKE ? OR lower(ifnull(basvuru_no,'')) LIKE ? OR lower(ifnull(yabanci_isim,'')) LIKE ? OR lower(ifnull(kimlik_no,'')) LIKE ? OR lower(ifnull(ulke,'')) LIKE ?)"
            params += [f"%{q}%"]*5
        if status_list:
            sql += " AND durum IN (" + ",".join("?"*len(status_list)) + ")"
            params += status_list
        sql += " ORDER BY (tarih IS NULL), tarih DESC, id DESC"
        rows = cur.execute(sql, params).fetchall()

    elif table == "ikamet_izni":
        sql = "SELECT * FROM ikamet_izni WHERE 1=1"; params=[]
        if q:
            sql += " AND (lower(ifnull(ad_soyad,'')) LIKE ? OR lower(ifnull(referans,'')) LIKE ? OR lower(ifnull(iletisim,'')) LIKE ? OR lower(ifnull(pasaport_tc,'')) LIKE ?)"
            params += [f"%{q}%"]*4
        if status_list:
            sql += " AND durum IN (" + ",".join("?"*len(status_list)) + ")"
            params += status_list
        sql += " ORDER BY (tarih IS NULL), tarih DESC, id DESC"
        rows = cur.execute(sql, params).fetchall()
    else:
        rows = cur.execute(f"SELECT * FROM {table} ORDER BY (tarih IS NULL), tarih DESC, id DESC").fetchall()

    conn.close(); return jsonify([dict(r) for r in rows])

@app.route("/api/save/<table>", methods=["POST"])
def save_row(table):
    allowed = {"calisma_izni","ikamet_izni","sgk","eimza"}
    if table not in allowed: return jsonify({"error":"invalid"}), 400
    data = request.json or {}
    conn = db(); cur = conn.cursor()
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row_id = data.get("id")
    tarih = parse_date_any(data.get("tarih"))

    if table == "calisma_izni":
        payload = (tarih, data.get("firma",""), data.get("basvuru_no",""), data.get("yabanci_isim",""),
                   data.get("kimlik_no",""), data.get("ulke",""), data.get("durum","DEĞERLENDİRME"), data.get("notlar",""))
        if row_id:
            cur.execute("UPDATE calisma_izni SET tarih=?,firma=?,basvuru_no=?,yabanci_isim=?,kimlik_no=?,ulke=?,durum=?,notlar=? WHERE id=?", payload+(row_id,))
        else:
            cur.execute("INSERT INTO calisma_izni(tarih,firma,basvuru_no,yabanci_isim,kimlik_no,ulke,durum,notlar,created_at) VALUES(?,?,?,?,?,?,?,?,?)", payload+(now,))
    elif table == "ikamet_izni":
        payload = (tarih, data.get("ad_soyad",""), data.get("referans",""), data.get("iletisim",""),
                   data.get("pasaport_tc",""), data.get("durum","DEĞERLENDİRME"), data.get("notlar",""))
        if row_id:
            cur.execute("UPDATE ikamet_izni SET tarih=?,ad_soyad=?,referans=?,iletisim=?,pasaport_tc=?,durum=?,notlar=? WHERE id=?", payload+(row_id,))
        else:
            cur.execute("INSERT INTO ikamet_izni(tarih,ad_soyad,referans,iletisim,pasaport_tc,durum,notlar,created_at) VALUES(?,?,?,?,?,?,?,?)", payload+(now,))
    elif table == "sgk":
        payload = (tarih, data.get("isveren_isim",""), data.get("tc",""), data.get("dosya_no",""), data.get("sistem_sifre",""),
                   data.get("isyeri_sifre",""), data.get("notlar",""))
        if row_id:
            cur.execute("UPDATE sgk SET tarih=?,isveren_isim=?,tc=?,dosya_no=?,sistem_sifre=?,isyeri_sifre=?,notlar=? WHERE id=?", payload+(row_id,))
        else:
            cur.execute("INSERT INTO sgk(tarih,isveren_isim,tc,dosya_no,sistem_sifre,isyeri_sifre,notlar,created_at) VALUES(?,?,?,?,?,?,?,?)", payload+(now,))
    else:
        payload = (tarih, data.get("musteri_adi",""), data.get("suresi",""), data.get("odeme",""),
                   data.get("kargo_durumu",""), data.get("notlar",""))
        if row_id:
            cur.execute("UPDATE eimza SET tarih=?,musteri_adi=?,suresi=?,odeme=?,kargo_durumu=?,notlar=? WHERE id=?", payload+(row_id,))
        else:
            cur.execute("INSERT INTO eimza(tarih,musteri_adi,suresi,odeme,kargo_durumu,notlar,created_at) VALUES(?,?,?,?,?,?,?)", payload+(now,))
    conn.commit(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/delete/<table>/<int:row_id>", methods=["POST"])
def delete_row(table, row_id):
    if table not in {"calisma_izni","ikamet_izni","sgk","eimza"}: return jsonify({"error":"invalid"}), 400
    conn=db(); cur=conn.cursor()
    cur.execute(f"DELETE FROM {table} WHERE id=?", (row_id,))
    conn.commit(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/import/<target>", methods=["POST"])
def api_import(target):
    if target not in {"calisma_izni","ikamet_izni","sgk","eimza"}: return jsonify({"error":"invalid"}), 400
    if "file" not in request.files: return jsonify({"error":"file missing"}), 400
    inserted = import_xlsx(request.files["file"], target)
    return jsonify({"inserted": inserted})

@app.route("/api/export/<table>")
def api_export(table):
    if table not in {"calisma_izni","ikamet_izni","sgk","eimza"}: return jsonify({"error":"invalid"}), 400
    bio = excel_export(table)
    name = {"calisma_izni":"calisma_izni.xlsx","ikamet_izni":"ikamet_izni.xlsx","sgk":"sgk.xlsx","eimza":"eimza.xlsx"}[table]
    return send_file(bio, as_attachment=True, download_name=name, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/api/backup/export")
def backup_export():
    conn=db(); cur=conn.cursor()
    data={}
    for t in ("calisma_izni","ikamet_izni","sgk","eimza"):
        data[t]=[dict(r) for r in cur.execute(f"SELECT * FROM {t}").fetchall()]
    conn.close()
    bio=io.BytesIO(json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8"))
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name="yedek.json", mimetype="application/json")

@app.route("/api/backup/import", methods=["POST"])
def backup_import():
    if "file" not in request.files: return jsonify({"error":"file missing"}), 400
    raw = request.files["file"].read()
    data = json.loads(raw.decode("utf-8"))
    conn=db(); cur=conn.cursor()
    for t in ("calisma_izni","ikamet_izni","sgk","eimza"): cur.execute(f"DELETE FROM {t}")
    for r in data.get("calisma_izni", []):
        cur.execute("INSERT INTO calisma_izni(id,tarih,firma,basvuru_no,yabanci_isim,kimlik_no,ulke,durum,notlar,created_at) VALUES(?,?,?,?,?,?,?,?,?,?)",
                    (r.get("id"),r.get("tarih"),r.get("firma"),r.get("basvuru_no"),r.get("yabanci_isim"),r.get("kimlik_no"),r.get("ulke"),r.get("durum"),r.get("notlar"),r.get("created_at")))
    for r in data.get("ikamet_izni", []):
        cur.execute("INSERT INTO ikamet_izni(id,tarih,ad_soyad,referans,iletisim,pasaport_tc,durum,notlar,created_at) VALUES(?,?,?,?,?,?,?,?,?)",
                    (r.get("id"),r.get("tarih"),r.get("ad_soyad"),r.get("referans"),r.get("iletisim"),r.get("pasaport_tc"),r.get("durum"),r.get("notlar"),r.get("created_at")))
    for r in data.get("sgk", []):
        cur.execute("INSERT INTO sgk(id,tarih,isveren_isim,tc,dosya_no,sistem_sifre,isyeri_sifre,notlar,created_at) VALUES(?,?,?,?,?,?,?,?)",
                    (r.get("id"),r.get("tarih"),r.get("isveren_isim"),r.get("tc"),r.get("sistem_sifre"),r.get("isyeri_sifre"),r.get("notlar"),r.get("created_at")))
    for r in data.get("eimza", []):
        cur.execute("INSERT INTO eimza(id,tarih,musteri_adi,suresi,odeme,kargo_durumu,notlar,created_at) VALUES(?,?,?,?,?,?,?,?)",
                    (r.get("id"),r.get("tarih"),r.get("musteri_adi"),r.get("suresi"),r.get("odeme"),r.get("kargo_durumu"),r.get("notlar"),r.get("created_at")))
    conn.commit(); conn.close()
    return jsonify({"ok": True})

@app.route("/api/reminders")
def reminders():
    c=db();cur=c.cursor()
    r=cur.execute("SELECT * FROM reminders ORDER BY (tarih IS NULL), tarih ASC").fetchall()
    c.close()
    return jsonify([dict(x) for x in r])

@app.route("/api/reminders/save",methods=["POST"])
def reminders_save():
    d=request.json or {}
    rid=d.get("id")
    tarih=parse_date_any(d.get("tarih"))
    baslik=d.get("baslik")
    aciklama=d.get("aciklama")
    now=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    c=db();cur=c.cursor()
    if rid:
        cur.execute("UPDATE reminders SET tarih=?,baslik=?,aciklama=? WHERE id=?",(tarih,baslik,aciklama,rid))
    else:
        cur.execute("INSERT INTO reminders(tarih,baslik,aciklama,created_at) VALUES(?,?,?,?)",(tarih,baslik,aciklama,now))
    c.commit();c.close()
    return {"ok":True}

@app.route("/api/reminders/delete/<int:i>",methods=["POST"])
def reminders_delete(i):
    c=db();cur=c.cursor()
    cur.execute("DELETE FROM reminders WHERE id=?",(i,))
    c.commit();c.close()
    return {"ok":True}

if __name__ == "__main__":
    init_db()
    app.run(host="0.0.0.0", port=5055, debug=False)